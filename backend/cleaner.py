"""
SAP AR Ledger Cleaning Pipeline — Phase 1
Pure function: clean_sap_books(file_bytes, ...) → (clean_df, sgl_v_df, CleaningReport)

Column reference (0-based positional index — NEVER use header name detection):
  col[4]  Clearing Document    ← P1: used for group matching
  col[5]  Document Type        ← GATE 1
  col[6]  Document Date        ← DATE field + FY date-range filter
  col[8]  Special G/L ind.     ← GATE 2
  col[10] Amount in local currency  ← AMOUNT
  col[14] Invoice reference    ← REF (PRIMARY identifier)
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from config import NOISE_THRESHOLD, date_to_fy_label
from pydantic import BaseModel


class CleaningReport(BaseModel):
    total_rows_input: int
    rows_after_cleaning: int
    excluded_null: int
    excluded_negative: int
    excluded_noise: int
    excluded_doc_type: int
    excluded_sgl: int
    excluded_date_fy: int = 0
    flagged_advance: int
    flagged_ab: int
    flagged_other_sgl: int
    duplicates_removed: int
    split_invoices_flagged: int
    used_fallback_doc_types: bool = False

logger = logging.getLogger(__name__)

# ── Document Type Rules ──────────────────────────────────────────────────────
PRIMARY_DOC_TYPES  = {"RV", "DR"}
EXCLUDE_DOC_TYPES  = {"CC", "BR"}

# ── Special G/L Indicator Rules ──────────────────────────────────────────────
EXCLUDE_SGL = {"L", "E", "U"}
FLAG_SGL = {
    "V": "SGL_V",
    "O": "SGL_O",
    "A": "SGL_A",
    "N": "SGL_N",
}


@dataclass
class _ExclusionCounters:
    null: int = 0
    negative: int = 0
    noise: int = 0
    doc_type: int = 0
    sgl: int = 0
    sgl_v_excluded: int = 0
    date_out_of_fy: int = 0
    dupe: int = 0
    flagged_advance: int = 0
    flagged_other_sgl: int = 0
    split_invoices: int = 0


def _parse_raw_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return pd.to_datetime(str(val), dayfirst=True).date()
    except Exception:
        return None


def _fmt_date(val: Any) -> Optional[str]:
    d = _parse_raw_date(val)
    return d.strftime("%d-%b-%Y") if d else (str(val) if val else None)


def _build_clean_df(
    raw_rows: List[Dict],
    c: _ExclusionCounters,
) -> pd.DataFrame:
    df = pd.DataFrame(raw_rows)
    if df.empty:
        return df

    final_rows: List[Dict] = []
    for ref, ref_group in df.groupby("invoice_ref"):
        if ref == "":
            # No invoice ref — deduplication not possible, keep all rows
            for _, r in ref_group.iterrows():
                final_rows.append(r.to_dict())
            continue

        # Sub-group by clearing_doc: the same invoice can be settled in
        # multiple separate clearing events (e.g., partial payments, two
        # different payment runs). These are DISTINCT entries, not duplicates.
        # True duplicate = same invoice_ref + same clearing_doc + same amount.
        for clr_doc, clr_group in ref_group.groupby("clearing_doc", sort=False):
            if clr_doc == "":
                # No clearing doc — cannot identify payment event, keep all
                for _, r in clr_group.iterrows():
                    final_rows.append(r.to_dict())
                continue

            unique_amounts = clr_group["amount"].unique()
            if len(unique_amounts) == 1:
                # Same invoice + same clearing doc + same amount → true duplicate
                c.dupe += len(clr_group) - 1
                final_rows.append(clr_group.iloc[0].to_dict())
            else:
                # Same invoice + same clearing doc + different amounts → partial clearing
                c.split_invoices += 1
                for _, r in clr_group.iterrows():
                    rd = r.to_dict()
                    ef = rd.get("flag", "")
                    rd["flag"] = f"{ef},SPLIT_INVOICE".strip(",") if ef else "SPLIT_INVOICE"
                    final_rows.append(rd)

    return pd.DataFrame(final_rows).reset_index(drop=True)


def clean_sap_books(
    file_bytes: bytes,
    noise_threshold: float = NOISE_THRESHOLD,
    fy_start: Optional[date] = None,
    fy_end: Optional[date] = None,
    doc_types_include: Optional[set] = None,
    doc_types_exclude: Optional[set] = None,
    exclude_sgl_v: bool = True,
    credit_note_handling_enabled: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame, CleaningReport]:
    """
    Parse and clean a raw SAP AR ledger Excel file.

    Parameters
    ----------
    file_bytes        : Raw .xlsx bytes
    noise_threshold   : Rows below this amount are excluded as noise
    fy_start          : P2: If provided, exclude SAP rows with doc_date < fy_start
    fy_end            : P2: If provided, exclude SAP rows with doc_date > fy_end
    doc_types_include : Override PRIMARY_DOC_TYPES (e.g. {"RV", "DC", "DR"})
    doc_types_exclude : Override EXCLUDE_DOC_TYPES (e.g. {"CC", "BR"})
    exclude_sgl_v     : If True, SGL_V entries go to separate sgl_v_df pool
                        instead of being included in clean_df

    Returns
    -------
    clean_df        : Main cleaned DataFrame for matching
    sgl_v_df        : Separate DataFrame of SGL_V (advance) entries (when exclude_sgl_v=True)
    report          : CleaningReport
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    # ── SAP header structure validation ──────────────────────────────────────
    # Verify the file has enough columns and the header row looks reasonable.
    # SAP columns are positional (not by name), so we validate structure, not labels.
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is not None:
        col_count = len([c for c in header_row if c is not None])
        if col_count < 10:
            logger.warning(
                "SAP file has only %d non-null header columns (expected ≥15). "
                "Columns are positional — ensure this is a raw SAP AR Ledger export, "
                "not a pre-processed workings file.",
                col_count,
            )

    # Use configurable doc types or fall back to defaults
    primary_doc_types = set(doc_types_include) if doc_types_include else PRIMARY_DOC_TYPES
    exclude_doc_types_set = set(doc_types_exclude) if doc_types_exclude else EXCLUDE_DOC_TYPES

    c = _ExclusionCounters()
    primary_rows: List[Dict] = []
    fallback_rows: List[Dict] = []
    sgl_v_rows: List[Dict] = []
    total_input = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_input += 1

        # ── Null amount ─────────────────────────────────────────────────────
        raw_amount = row[10]
        if raw_amount is None:
            c.null += 1
            continue
        try:
            amount = float(raw_amount)
        except (ValueError, TypeError):
            c.null += 1
            continue

        # ── Non-positive ────────────────────────────────────────────────────
        is_credit_note = False
        if amount < 0:
            if credit_note_handling_enabled:
                is_credit_note = True
                amount = abs(amount)  # Store as positive, flag as credit note
            else:
                c.negative += 1
                continue
        elif amount == 0:
            c.negative += 1
            continue

        # ── Noise filter ────────────────────────────────────────────────────
        if amount < noise_threshold:
            c.noise += 1
            continue

        # ── Document Type gate ──────────────────────────────────────────────
        doc_type = str(row[5]).strip() if row[5] is not None else ""
        if doc_type in exclude_doc_types_set:
            c.doc_type += 1
            continue
        is_primary = doc_type in primary_doc_types

        # ── P2: SAP date window ─────────────────────────────────────────────
        doc_date_raw = row[6]
        doc_date = _parse_raw_date(doc_date_raw)
        if fy_start and fy_end and doc_date is not None:
            if not (fy_start <= doc_date <= fy_end):
                c.date_out_of_fy += 1
                continue
        # If doc_date is None but we have a valid amount, KEEP the row
        # (don't lose matching candidates just because date is unparseable)

        # ── Special G/L Indicator gate ──────────────────────────────────────
        sgl = str(row[8]).strip() if row[8] is not None else ""
        flag = ""
        if sgl in EXCLUDE_SGL:
            c.sgl += 1
            continue
        elif sgl in FLAG_SGL:
            flag = FLAG_SGL[sgl]
            if sgl == "V":
                c.flagged_advance += 1
            else:
                c.flagged_other_sgl += 1

        # ── P1: Extract Clearing Document (col[4]) ─────────────────────────
        clearing_doc = str(row[4]).strip() if row[4] is not None else ""

        # ── P4: Compute SAP FY from doc_date ────────────────────────────────
        sap_fy = date_to_fy_label(doc_date) if doc_date else ""

        # ── Credit note flag ──────────────────────────────────────────────
        if is_credit_note:
            flag = f"{flag},CREDIT_NOTE".strip(",") if flag else "CREDIT_NOTE"

        # ── Collect ─────────────────────────────────────────────────────────
        record = {
            "doc_date":      _fmt_date(doc_date_raw),
            "amount":        amount,
            "invoice_ref":   str(row[14]).strip() if row[14] is not None else "",
            "doc_type":      doc_type,
            "sgl_ind":       sgl,
            "flag":          flag,
            "clearing_doc":  clearing_doc,
            "sap_fy":        sap_fy,
        }

        # ── SGL_V separation (advance entries) ──────────────────────────────
        if sgl == "V" and exclude_sgl_v:
            sgl_v_rows.append(record)
            c.sgl_v_excluded += 1
            continue

        if is_primary:
            primary_rows.append(record)
        else:
            fallback_rows.append(record)

    wb.close()

    used_fallback = False
    if primary_rows:
        rows_to_use = primary_rows
        c.doc_type += len(fallback_rows)
    else:
        rows_to_use = fallback_rows
        used_fallback = True
        logger.warning(
            "No %s rows found — falling back to all %d valid rows",
            "/".join(sorted(primary_doc_types)),
            len(fallback_rows),
        )

    clean_df = _build_clean_df(rows_to_use, c)

    if used_fallback and not clean_df.empty:
        clean_df["flag"] = clean_df["flag"].apply(
            lambda f: f"{f},FALLBACK_DOCTYPE".strip(",") if f else "FALLBACK_DOCTYPE"
        )

    # Build separate SGL_V DataFrame
    sgl_v_df = pd.DataFrame(sgl_v_rows).reset_index(drop=True) if sgl_v_rows else pd.DataFrame()

    report = CleaningReport(
        total_rows_input=total_input,
        rows_after_cleaning=len(clean_df),
        excluded_null=c.null,
        excluded_negative=c.negative,
        excluded_noise=c.noise,
        excluded_doc_type=c.doc_type,
        excluded_sgl=c.sgl,
        sgl_v_excluded=c.sgl_v_excluded,
        excluded_date_fy=c.date_out_of_fy,
        flagged_advance=c.flagged_advance,
        flagged_ab=0,
        flagged_other_sgl=c.flagged_other_sgl,
        duplicates_removed=c.dupe,
        split_invoices_flagged=c.split_invoices,
        used_fallback_doc_types=used_fallback,
    )

    logger.info(
        "SAP cleaning: %d raw → %d clean (%d SGL_V separated) | "
        "excluded: null=%d neg=%d noise=%d doctype=%d sgl=%d sgl_v=%d fy=%d dupes=%d | fallback=%s",
        total_input, len(clean_df), len(sgl_v_df),
        c.null, c.negative, c.noise, c.doc_type, c.sgl, c.sgl_v_excluded, c.date_out_of_fy, c.dupe,
        used_fallback,
    )
    return clean_df, sgl_v_df, report
