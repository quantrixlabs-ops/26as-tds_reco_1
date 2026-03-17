"""
Excel Output Generator — Phase 1 v2
generate_excel(reco_result, cleaning_report, fy_label) → bytes

5-sheet workbook:
  Sheet 1 — Summary (with confidence tier counts + cross-FY stats)
  Sheet 2 — Matched Pairs (Invoice Ref as PRIMARY column, confidence, clearing doc, SAP FY)
  Sheet 3 — Unmatched 26AS (with best candidate info + rejection reason)
  Sheet 4 — Unmatched Books
  Sheet 5 — Variance Analysis
"""
from __future__ import annotations

import io
import statistics
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from config import DEFAULT_FINANCIAL_YEAR, VARIANCE_CAP_PCT
from models import CleaningReport, MatchedPair, RecoResult, UnmatchedAs26Entry

# ── Colour constants ──────────────────────────────────────────────────────────
NAVY       = "1F3864"
WHITE      = "FFFFFF"
LIGHT_BLUE = "EBF3FB"

VAR_GREEN  = "C6EFCE"
VAR_YELLOW = "FFEB9C"
VAR_RED    = "FFC7CE"

VIOLATION_RED = "FF0000"
GOOD_GREEN    = "00B050"
CONF_HIGH     = "C6EFCE"   # green
CONF_MEDIUM   = "FFEB9C"   # yellow

INR_FMT = '#,##0.00'


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=WHITE, size=10, name="Arial") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_style(ws, row, start_col, end_col, value):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    cell.fill = _fill(NAVY)
    cell.font = _font(bold=True, color=WHITE, size=11)
    cell.alignment = _align(h="center")


def _col_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(NAVY)
    cell.font = _font(bold=True, color=WHITE, size=10)
    cell.alignment = _align(h="center")
    cell.border = _border()


def _data_cell(ws, row, col, value, fmt=None, bg=None, bold=False, align_h="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _border()
    cell.alignment = _align(h=align_h)
    if fmt:
        cell.number_format = fmt
    bg_color = bg if bg else (LIGHT_BLUE if row % 2 == 0 else WHITE)
    cell.fill = _fill(bg_color)
    cell.font = Font(bold=bold, size=10, name="Arial", color="000000")


def _autofit(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _var_color(pct):
    if abs(pct) < 1:
        return VAR_GREEN
    elif abs(pct) <= 5:
        return VAR_YELLOW
    else:
        return VAR_RED


def _conf_color(conf):
    if conf == "HIGH":
        return CONF_HIGH
    elif conf == "MEDIUM":
        return CONF_MEDIUM
    return VAR_RED


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _build_summary(ws, result: RecoResult, report: CleaningReport, fy_label: str):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    now_str = datetime.now().strftime("%d-%b-%Y %H:%M")
    title = f"{result.deductor_name} | TDS Reconciliation | {fy_label} | Generated: {now_str}"
    _header_style(ws, 1, 1, 4, title)

    row = 3
    ws.cell(row=row, column=1, value="CLEANING STATISTICS").font = _font(bold=True, color="000000", size=10)
    row += 1

    cleaning_rows = [
        ("Total raw rows (input)",              report.total_rows_input),
        ("Rows after cleaning",                 report.rows_after_cleaning),
        ("  Excluded — null amount",            report.excluded_null),
        ("  Excluded — negative/zero",          report.excluded_negative),
        ("  Excluded — noise (<₹100)",          report.excluded_noise),
        ("  Excluded — doc type (CC/BR/other)", report.excluded_doc_type),
        ("  Excluded — Special G/L (L/E/U)",    report.excluded_sgl),
        ("  Excluded — outside SAP date window", report.excluded_date_fy),
        ("  Flagged — advance (SGL=V)",         report.flagged_advance),
        ("  Flagged — other SGL (O/A/N)",       report.flagged_other_sgl),
        ("  Duplicates removed",                report.duplicates_removed),
        ("  Split invoices flagged",            report.split_invoices_flagged),
    ]
    for label, value in cleaning_rows:
        _data_cell(ws, row, 1, label)
        _data_cell(ws, row, 2, value, align_h="right")
        row += 1

    if report.used_fallback_doc_types:
        _data_cell(ws, row, 1, "⚠ No RV/DC/DR rows — fallback doc types used", bg="FFEB9C")
        _data_cell(ws, row, 2, "FALLBACK", align_h="center", bg="FFEB9C")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="RECONCILIATION STATISTICS").font = _font(bold=True, color="000000", size=10)
    row += 1

    match_color = (
        VAR_GREEN  if result.match_rate_pct == 100 else
        VAR_YELLOW if result.match_rate_pct >= 80  else VAR_RED
    )

    reco_rows = [
        ("Financial Year",          fy_label,                              None),
        ("Deductor Name",           result.deductor_name,                  None),
        ("TAN",                     result.tan,                            None),
        ("Name Match Score",
            f"{result.fuzzy_score:.1f}%" if result.fuzzy_score else "Manual override", None),
        ("Total 26AS entries",      result.total_26as_entries,             None),
        ("Matched entries",         result.matched_count,                  None),
        ("Match rate",              f"{result.match_rate_pct:.2f}%",       match_color),
        (f"  HIGH confidence (≤1%)", result.high_confidence_count,         CONF_HIGH),
        (f"  MEDIUM confidence (1–{VARIANCE_CAP_PCT:.0f}%)",
            result.medium_confidence_count, CONF_MEDIUM),
        ("Unmatched 26AS entries",  result.unmatched_26as_count,           None),
        ("Unmatched book invoices", result.unmatched_books_count,          None),
        ("Average variance %",      f"{result.avg_variance_pct:.2f}%",    None),
        ("Cross-FY matches",        result.cross_fy_match_count,          None),
        ("Constraint violations",   result.constraint_violations,
            VIOLATION_RED if result.constraint_violations > 0 else VAR_GREEN),
        (f"Variance cap applied",   f"{VARIANCE_CAP_PCT:.0f}%",           None),
    ]

    for label, value, bg in reco_rows:
        _data_cell(ws, row, 1, label)
        _data_cell(ws, row, 2, value, align_h="right", bg=bg)
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30


# ── Sheet 2: Matched Pairs ───────────────────────────────────────────────────

def _build_matched(ws, pairs: List[MatchedPair]):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # P4: Invoice Ref is the PRIMARY column — placed first after #
    headers = [
        "#",
        "Invoice Ref(s)",       # PRIMARY identifier
        "26AS Date",
        "26AS Amount (₹)",
        "Section",
        "Books Sum (₹)",
        "Variance (₹)",
        "Variance %",
        "Confidence",           # P3
        "Match Type",
        "Invoice Count",
        "Invoice Date(s)",
        "Invoice Amounts",
        "SGL Flags",
        "Clearing Doc(s)",      # P1/P4
        "SAP FY",               # P4
        "Cross-FY",             # P4
    ]
    for c, h in enumerate(headers, 1):
        _col_header(ws, 1, c, h)

    sorted_pairs = sorted(pairs, key=lambda p: p.as26_amount, reverse=True)

    for r, pair in enumerate(sorted_pairs, 2):
        var_bg = _var_color(pair.variance_pct)
        conf_bg = _conf_color(pair.confidence)

        _data_cell(ws, r, 1,  r - 1,                                          align_h="center")
        _data_cell(ws, r, 2,  ", ".join(pair.invoice_refs), bold=True)           # PRIMARY
        _data_cell(ws, r, 3,  pair.as26_date or "",                            align_h="center")
        _data_cell(ws, r, 4,  pair.as26_amount,    fmt=INR_FMT,               align_h="right")
        _data_cell(ws, r, 5,  pair.section,                                    align_h="center")
        _data_cell(ws, r, 6,  pair.books_sum,      fmt=INR_FMT,               align_h="right")
        _data_cell(ws, r, 7,  pair.variance_amt,   fmt=INR_FMT, bg=var_bg,    align_h="right")
        _data_cell(ws, r, 8,  f"{pair.variance_pct:.2f}%",      bg=var_bg,    align_h="right")
        _data_cell(ws, r, 9,  pair.confidence,                   bg=conf_bg,   align_h="center")
        _data_cell(ws, r, 10, pair.match_type,                                 align_h="center")
        _data_cell(ws, r, 11, pair.invoice_count,                              align_h="center")
        _data_cell(ws, r, 12, ", ".join(d or "" for d in pair.invoice_dates))
        _data_cell(ws, r, 13, ", ".join(f"{a:,.2f}" for a in pair.invoice_amounts))
        _data_cell(ws, r, 14, ", ".join(f for f in pair.sgl_flags if f))
        _data_cell(ws, r, 15, ", ".join(pair.clearing_docs) if pair.clearing_docs else "")
        _data_cell(ws, r, 16, ", ".join(set(pair.sap_fys)) if pair.sap_fys else "")
        _data_cell(ws, r, 17, "YES" if pair.cross_fy else "",
                   bg="FFEB9C" if pair.cross_fy else None,                      align_h="center")

    _autofit(ws)


# ── Sheet 3: Unmatched 26AS ──────────────────────────────────────────────────

def _build_unmatched_26as(ws, entries: List[UnmatchedAs26Entry]):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "#", "26AS Date", "Amount (₹)", "Section", "TAN",
        "Best Candidate Ref", "Best Candidate Amt (₹)",
        "Best Candidate Var %", "Rejection Reason",
    ]
    for c, h in enumerate(headers, 1):
        _col_header(ws, 1, c, h)

    if not entries:
        ws.merge_cells("A2:I2")
        cell = ws.cell(row=2, column=1, value="✓ All 26AS entries matched")
        cell.fill = _fill(VAR_GREEN)
        cell.font = Font(bold=True, color="375623", size=11, name="Arial")
        cell.alignment = _align(h="center")
        _autofit(ws)
        return

    for r, entry in enumerate(entries, 2):
        _data_cell(ws, r, 1, r - 1,                             align_h="center")
        _data_cell(ws, r, 2, entry.transaction_date or "",       align_h="center")
        _data_cell(ws, r, 3, entry.amount, fmt=INR_FMT,         align_h="right")
        _data_cell(ws, r, 4, entry.section,                      align_h="center")
        _data_cell(ws, r, 5, entry.tan,                          align_h="center")
        _data_cell(ws, r, 6, entry.best_candidate_ref or "—")
        _data_cell(ws, r, 7, entry.best_candidate_amount, fmt=INR_FMT,
                   align_h="right") if entry.best_candidate_amount else _data_cell(ws, r, 7, "—")
        _data_cell(ws, r, 8,
                   f"{entry.best_candidate_variance_pct:.2f}%"
                   if entry.best_candidate_variance_pct is not None else "—",
                   bg=VAR_RED if entry.best_candidate_variance_pct and
                   abs(entry.best_candidate_variance_pct) > VARIANCE_CAP_PCT else None,
                   align_h="right")
        _data_cell(ws, r, 9, entry.rejection_reason)

    _autofit(ws)


# ── Sheet 4: Unmatched Books ─────────────────────────────────────────────────

def _possible_reason_books(entry) -> str:
    if "SGL_V" in (entry.flag or ""):
        return "Advance payment — TDS may be on advance"
    if entry.amount > 1_000_000:
        return "Large milestone / different financial year"
    return "Timing difference — may appear in 26AS next period"


def _build_unmatched_books(ws, entries):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "#", "Invoice Ref", "Invoice Date", "Amount (₹)",
        "Doc Type", "SGL Flag", "Clearing Doc", "SAP FY", "Possible Reason",
    ]
    for c, h in enumerate(headers, 1):
        _col_header(ws, 1, c, h)

    if not entries:
        ws.merge_cells("A2:I2")
        cell = ws.cell(row=2, column=1, value="✓ All book invoices matched")
        cell.fill = _fill(VAR_GREEN)
        cell.font = Font(bold=True, color="375623", size=11, name="Arial")
        cell.alignment = _align(h="center")
        _autofit(ws)
        return

    sorted_entries = sorted(entries, key=lambda e: e.amount, reverse=True)
    for r, entry in enumerate(sorted_entries, 2):
        _data_cell(ws, r, 1, r - 1,                      align_h="center")
        _data_cell(ws, r, 2, entry.invoice_ref, bold=True)  # PRIMARY
        _data_cell(ws, r, 3, entry.doc_date or "",        align_h="center")
        _data_cell(ws, r, 4, entry.amount, fmt=INR_FMT,  align_h="right")
        _data_cell(ws, r, 5, entry.doc_type,               align_h="center")
        _data_cell(ws, r, 6, entry.flag or "",             align_h="center")
        _data_cell(ws, r, 7, entry.clearing_doc or "")
        _data_cell(ws, r, 8, entry.sap_fy or "",           align_h="center")
        _data_cell(ws, r, 9, _possible_reason_books(entry))

    _autofit(ws)


# ── Sheet 5: Variance Analysis ───────────────────────────────────────────────

def _build_variance(ws, pairs: List[MatchedPair]):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Bucket", "Count", "% of Matched", "Interpretation"]
    for c, h in enumerate(headers, 1):
        _col_header(ws, 1, c, h)

    buckets = [
        ("0% Exact",  lambda p: abs(p.variance_pct) < 0.01,     VAR_GREEN,  "Perfect match"),
        ("0–1%",      lambda p: 0.01 <= abs(p.variance_pct) < 1, VAR_GREEN,  "HIGH confidence"),
        ("1–5%",      lambda p: 1 <= abs(p.variance_pct) < 5,    VAR_YELLOW, "MEDIUM confidence — acceptable"),
        ("5–10%",     lambda p: 5 <= abs(p.variance_pct) < 10,   VAR_RED,    "Rejected by variance cap"),
        ("10–20%",    lambda p: 10 <= abs(p.variance_pct) < 20,  VAR_RED,    "Rejected by variance cap"),
        ("20–50%",    lambda p: 20 <= abs(p.variance_pct) < 50,  VAR_RED,    "Rejected by variance cap"),
        (">50%",      lambda p: abs(p.variance_pct) >= 50,        VAR_RED,    "Rejected by variance cap"),
    ]

    total = len(pairs)
    for r, (label, pred, color, interp) in enumerate(buckets, 2):
        count = sum(1 for p in pairs if pred(p))
        pct = (count / total * 100) if total > 0 else 0.0
        _data_cell(ws, r, 1, label,          bg=color)
        _data_cell(ws, r, 2, count,          bg=color, align_h="center")
        _data_cell(ws, r, 3, f"{pct:.1f}%",  bg=color, align_h="center")
        _data_cell(ws, r, 4, interp,         bg=color)

    if pairs:
        variances = [p.variance_pct for p in pairs]
        stats_row = len(buckets) + 3
        ws.cell(row=stats_row, column=1, value="SUMMARY STATISTICS").font = (
            _font(bold=True, color="000000", size=10)
        )
        stats_row += 1
        for label, value in [
            ("Min variance %",     f"{min(variances):.2f}%"),
            ("Max variance %",     f"{max(variances):.2f}%"),
            ("Average variance %", f"{sum(variances)/len(variances):.2f}%"),
            ("Median variance %",  f"{statistics.median(variances):.2f}%"),
            ("Std Dev %",          f"{statistics.stdev(variances):.2f}%" if len(variances) > 1 else "N/A"),
        ]:
            _data_cell(ws, stats_row, 1, label)
            _data_cell(ws, stats_row, 2, value, align_h="right")
            stats_row += 1

        # Confidence summary
        stats_row += 1
        ws.cell(row=stats_row, column=1, value="CONFIDENCE BREAKDOWN").font = (
            _font(bold=True, color="000000", size=10)
        )
        stats_row += 1
        high = sum(1 for p in pairs if p.confidence == "HIGH")
        med  = sum(1 for p in pairs if p.confidence == "MEDIUM")
        _data_cell(ws, stats_row, 1, "HIGH (≤1%)", bg=CONF_HIGH)
        _data_cell(ws, stats_row, 2, f"{high} ({high/total*100:.1f}%)", bg=CONF_HIGH, align_h="right")
        stats_row += 1
        _data_cell(ws, stats_row, 1, f"MEDIUM (1–{VARIANCE_CAP_PCT:.0f}%)", bg=CONF_MEDIUM)
        _data_cell(ws, stats_row, 2, f"{med} ({med/total*100:.1f}%)", bg=CONF_MEDIUM, align_h="right")

    _autofit(ws)


# ── Main generator ────────────────────────────────────────────────────────────

def generate_excel(
    result: RecoResult,
    report: CleaningReport,
    fy_label: str = DEFAULT_FINANCIAL_YEAR,
) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    _build_summary(ws1, result, report, fy_label)

    ws2 = wb.create_sheet("Matched Pairs")
    _build_matched(ws2, result.matched_pairs)

    ws3 = wb.create_sheet("Unmatched 26AS")
    _build_unmatched_26as(ws3, result.unmatched_26as)

    ws4 = wb.create_sheet("Unmatched Books")
    _build_unmatched_books(ws4, result.unmatched_books)

    ws5 = wb.create_sheet("Variance Analysis")
    _build_variance(ws5, result.matched_pairs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
