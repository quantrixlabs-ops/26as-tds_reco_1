"""
Excel Generator v2 — enterprise-grade output with full audit metadata.
Sheets: Summary (with run ID, hashes, algorithm version, control totals),
        Matched Pairs (with composite score breakdown),
        Suggested Matches (requires authorization),
        Exceptions (REQUIRES REVIEW),
        Unmatched 26AS / Unmatched Books / Variance Analysis.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db.models import ReconciliationRun, MatchedPair, Unmatched26AS, UnmatchedBook, ExceptionRecord, SuggestedMatch
from core.settings import settings

# Colour palette
NAVY      = "1B3A5C"
WHITE     = "FFFFFF"
LIGHT_BLUE= "D6E4F0"
VAR_GREEN = "D4EDDA"
VAR_YELLOW= "FFF3CD"
VAR_RED   = "F8D7DA"
CONF_HIGH = "C8F7C5"
CONF_MED  = "FFF9C4"
CONF_LOW  = "FFCCBC"
ORANGE    = "FF6B35"
AMBER     = "F59E0B"
PURPLE    = "7B2D8B"
GRAY_LIGHT= "F5F5F5"
GRAY_SEP  = "EEEEEE"
CRITICAL_RED = "C62828"

# Category colours for suggested matches
CAT_COLORS = {
    "FORCE": "E53935",
    "HIGH_VARIANCE_20_PLUS": "C62828",
    "HIGH_VARIANCE_3_20": "FB8C00",
    "DATE_SOFT_PREFERENCE": "1565C0",
    "ADVANCE_PAYMENT": "7B2D8B",
    "CROSS_FY": "D84315",
    "TIER_CAP_EXCEEDED": "6D4C41",
}


def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=WHITE, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_title_header_sep(ws, title: str, headers: list, merge_end_col: str, title_fill: str = NAVY):
    """Write a standard 3-row header block: title (row 1), headers (row 2), separator (row 3).

    Data should start at row 4.  Freeze pane is set to A4.
    """
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)

    # Row 1: merged title
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws["A1"].fill = _fill(title_fill)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 26

    # Row 2: column headers
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        cell.fill = _fill(NAVY)
        cell.alignment = _align("center")

    # Row 3: thin separator — prevents merge/freeze interaction that hides row 4 in some viewers
    ws.row_dimensions[3].height = 6
    for c in range(1, num_cols + 1):
        ws.cell(3, c).fill = _fill(GRAY_SEP)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


# ── Single-run Excel ─────────────────────────────────────────────────────────

def generate_excel_v2(
    run: ReconciliationRun,
    matched_pairs: List[MatchedPair],
    unmatched_26as: List[Unmatched26AS],
    unmatched_books: List[UnmatchedBook],
    exceptions: List[ExceptionRecord],
    suggested_matches: List[SuggestedMatch] | None = None,
    variance_thresholds: tuple[float, float] = (1.0, 3.0),
) -> bytes:
    # Deduplicate matched pairs by as26_row_hash (safety net)
    seen_hashes: set = set()
    deduped_pairs: List[MatchedPair] = []
    for mp in matched_pairs:
        h = mp.as26_row_hash
        if h and h in seen_hashes:
            continue
        if h:
            seen_hashes.add(h)
        deduped_pairs.append(mp)
    matched_pairs = deduped_pairs

    wb = Workbook()

    _build_summary(wb.active, run, matched_pairs, unmatched_26as, exceptions)
    wb.active.title = "Summary"

    ws_exc = wb.create_sheet("⚠ Requires Review")
    _build_exceptions(ws_exc, exceptions)

    var_green, var_yellow = variance_thresholds

    ws_match = wb.create_sheet("Matched Pairs")
    _build_matched(ws_match, matched_pairs, run, var_green=var_green, var_yellow=var_yellow)

    if suggested_matches:
        ws_sug = wb.create_sheet("Suggested Matches")
        _build_suggested(ws_sug, suggested_matches, run, var_green=var_green, var_yellow=var_yellow)

    ws_un26 = wb.create_sheet("Unmatched 26AS")
    _build_unmatched_26as(ws_un26, unmatched_26as)

    ws_unbks = wb.create_sheet("Unmatched SAP Books")
    _build_unmatched_books(ws_unbks, unmatched_books)

    if matched_pairs:
        ws_var = wb.create_sheet("Variance Analysis")
        _build_variance(ws_var, matched_pairs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_summary(ws, run: ReconciliationRun, matched, unmatched_26as, exceptions):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = f"TDS RECONCILIATION — {run.financial_year} | RUN-{run.run_number:04d}"
    ws["A1"].font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 30

    # Audit metadata block
    rows = [
        ("Run ID", run.id),
        ("Run Number", f"RUN-{run.run_number:04d}"),
        ("Financial Year", run.financial_year),
        ("Deductor", run.deductor_name or "—"),
        ("TAN", run.tan or "—"),
        ("Algorithm Version", run.algorithm_version),
        ("SAP File", run.sap_filename),
        ("SAP File Hash (SHA-256)", run.sap_file_hash),
        ("26AS File", run.as26_filename),
        ("26AS File Hash (SHA-256)", run.as26_file_hash),
        ("Output Hash (SHA-256)", run.output_hash or "Pending"),
        ("Generated At", datetime.now(timezone.utc).strftime("%d-%b-%Y %H:%M UTC")),
        ("Status", run.status),
        ("Reviewed By", run.reviewed_by_id or "Pending"),
    ]
    for i, (label, value) in enumerate(rows, 2):
        ws.cell(i, 1, label).font = Font(bold=True, size=9, name="Calibri", color="333333")
        ws.cell(i, 2, str(value)).font = Font(size=9, name="Calibri")
        ws.cell(i, 1).fill = _fill(GRAY_LIGHT)

    # Separator
    sep_row = len(rows) + 3
    ws.cell(sep_row, 1, "RECONCILIATION METRICS").font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws.cell(sep_row, 1).fill = _fill(NAVY)
    ws.merge_cells(f"A{sep_row}:H{sep_row}")
    ws.cell(sep_row, 1).alignment = _align("center")

    # Metrics
    total = run.total_26as_entries
    rate_color = VAR_GREEN if run.match_rate_pct >= 95 else VAR_YELLOW if run.match_rate_pct >= 75 else VAR_RED
    metrics = [
        ("Total 26AS Entries", total, None),
        ("Matched", run.matched_count, VAR_GREEN),
        ("Suggested (Requires Authorization)", run.suggested_count or 0,
         VAR_YELLOW if (run.suggested_count or 0) > 0 else VAR_GREEN),
        ("Unmatched 26AS", run.unmatched_26as_count, VAR_RED if run.unmatched_26as_count > 0 else VAR_GREEN),
        ("Match Rate", f"{run.match_rate_pct:.2f}%", rate_color),
        ("HIGH Confidence", run.high_confidence_count, CONF_HIGH),
        ("MEDIUM Confidence", run.medium_confidence_count, CONF_MED),
        ("LOW Confidence", run.low_confidence_count, CONF_LOW),
        ("Exceptions (Requires Review)", len(exceptions), VAR_RED if exceptions else VAR_GREEN),
        ("", "", None),
        ("CONTROL TOTALS", "", NAVY),
        ("Total 26AS Amount (₹)", f"{run.total_26as_amount:,.2f}", None),
        ("Matched Amount (₹)", f"{run.matched_amount:,.2f}", None),
        ("Unmatched Amount (₹)", f"{run.unmatched_26as_amount:,.2f}", None),
        ("Control Total Balanced", "✓ YES" if run.control_total_balanced else "✗ NO",
         VAR_GREEN if run.control_total_balanced else VAR_RED),
        ("", "", None),
        ("DATA QUALITY FLAGS", "", NAVY),
        ("PAN Issues", "⚠ YES" if run.has_pan_issues else "✓ None", VAR_RED if run.has_pan_issues else VAR_GREEN),
        ("Rate Mismatches", "⚠ YES" if run.has_rate_mismatches else "✓ None",
         VAR_RED if run.has_rate_mismatches else VAR_GREEN),
        ("Duplicate 26AS Entries", "⚠ YES" if run.has_duplicate_26as else "✓ None",
         VAR_RED if run.has_duplicate_26as else VAR_GREEN),
    ]

    for i, (label, value, bg) in enumerate(metrics, sep_row + 1):
        c1 = ws.cell(i, 1, label)
        c2 = ws.cell(i, 2, value)
        c1.font = Font(bold=bool(label and label == label.upper()), size=9, name="Calibri")
        c2.font = Font(size=9, name="Calibri", bold=bool(bg == NAVY))
        if bg == NAVY:
            c1.fill = _fill(NAVY)
            c1.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            ws.merge_cells(f"A{i}:H{i}")
            c1.alignment = _align("center")
        elif bg:
            c2.fill = _fill(bg)

    # Disclaimer
    disc_row = sep_row + len(metrics) + 3
    ws.merge_cells(f"A{disc_row}:H{disc_row}")
    ws.cell(disc_row, 1,
            "⚠ This reconciliation output is a working paper requiring CA review and sign-off "
            "before use in any audit, assessment, or litigation proceeding.").font = Font(
        italic=True, size=8, color="888888", name="Calibri")
    ws.cell(disc_row, 1).alignment = _align("left", wrap=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60


def _build_exceptions(ws, exceptions: List[ExceptionRecord]):
    headers = ["#", "Type", "Severity", "Amount (₹)", "Section", "Description", "Status"]
    _write_title_header_sep(
        ws,
        "⚠ REQUIRES REVIEW — All items below need mandatory CA review before sign-off",
        headers,
        get_column_letter(len(headers)),
        title_fill=ORANGE,
    )

    sev_colors = {"CRITICAL": "C62828", "HIGH": "E53935", "MEDIUM": "FB8C00", "LOW": "F9A825"}

    for r, exc in enumerate(exceptions, 4):
        sev_color = sev_colors.get(exc.severity, "888888")
        ws.cell(r, 1, r - 3).alignment = _align("center")
        ws.cell(r, 2, exc.exception_type.replace("_", " "))
        ws.cell(r, 3, exc.severity).fill = _fill(sev_color)
        ws.cell(r, 3).font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        ws.cell(r, 3).alignment = _align("center")
        ws.cell(r, 4, f"₹{exc.amount:,.2f}" if exc.amount else "—").alignment = _align("right")
        ws.cell(r, 5, exc.section or "—").alignment = _align("center")
        ws.cell(r, 6, exc.description).alignment = _align("left", wrap=True)
        status = "✓ Reviewed" if exc.reviewed else "⏳ Pending"
        ws.cell(r, 7, status).fill = _fill(VAR_GREEN if exc.reviewed else VAR_YELLOW)
        ws.row_dimensions[r].height = 40

    widths = [5, 22, 12, 16, 12, 70, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_matched(ws, pairs: List[MatchedPair], run: ReconciliationRun,
                   var_green: float = 1.0, var_yellow: float = 3.0):
    headers = [
        "#", "26AS Date", "Section", "26AS Amount (₹)", "Books Sum (₹)",
        "Variance ₹", "Variance %", "Match Type", "Confidence",
        "Composite Score", "Invoice Refs", "Clearing Doc",
        "Cross-FY", "AI Risk", "Prior Year", "Rate ⚠",
    ]
    title = (
        f"Matched Pairs — {run.deductor_name} | Run #{run.run_number} "
        f"| {run.financial_year} | Algorithm {run.algorithm_version}"
    )
    _write_title_header_sep(ws, title, headers, get_column_letter(len(headers)))

    conf_colors = {"HIGH": CONF_HIGH, "MEDIUM": CONF_MED, "LOW": CONF_LOW}

    for r, p in enumerate(pairs, 4):
        conf_color = conf_colors.get(p.confidence or "", "FFFFFF")
        var_pct = p.variance_pct or 0
        score = p.composite_score or 0
        var_color = VAR_GREEN if var_pct <= var_green else VAR_YELLOW if var_pct <= var_yellow else VAR_RED
        score_color = CONF_HIGH if score >= 80 else CONF_MED if score >= 60 else CONF_LOW

        ws.cell(r, 1, r - 3).alignment = _align("center")
        ws.cell(r, 2, p.as26_date or "—")
        ws.cell(r, 3, p.section or "—").alignment = _align("center")
        ws.cell(r, 4, p.as26_amount or 0).number_format = '#,##0.00'
        ws.cell(r, 5, p.books_sum or 0).number_format = '#,##0.00'
        ws.cell(r, 6, p.variance_amt or 0).fill = _fill(var_color); ws.cell(r, 6).number_format = '#,##0.00'
        ws.cell(r, 7, f"{var_pct:.2f}%").fill = _fill(var_color)
        ws.cell(r, 8, p.match_type or "—")
        ws.cell(r, 9, p.confidence or "—").fill = _fill(conf_color)
        ws.cell(r, 10, f"{score:.1f}").fill = _fill(score_color)
        ws.cell(r, 11, ", ".join(p.invoice_refs or []))
        ws.cell(r, 12, p.clearing_doc or "—")
        ws.cell(r, 13, "Y" if p.cross_fy else "").alignment = _align("center")
        ws.cell(r, 14, "!" if p.ai_risk_flag else "").alignment = _align("center")
        ws.cell(r, 15, "Y" if p.is_prior_year else "").alignment = _align("center")
        # Rate mismatch flag (Fix 5)
        rate_cell = ws.cell(r, 16, "⚠" if p.rate_mismatch else "")
        rate_cell.alignment = _align("center")
        if p.rate_mismatch:
            rate_cell.fill = _fill(VAR_RED)

    widths = [5, 14, 10, 16, 16, 14, 12, 18, 11, 14, 35, 16, 10, 9, 12, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_suggested(ws, suggestions: List[SuggestedMatch], run: ReconciliationRun,
                     var_green: float = 1.0, var_yellow: float = 3.0):
    """Build a 'Requires Authorization' sheet for suggested matches."""
    headers = [
        "#", "Category", "26AS Date", "Section", "26AS Amount (₹)", "Books Sum (₹)",
        "Variance ₹", "Variance %", "Match Type", "Confidence", "Composite Score",
        "Invoice Refs", "Status", "Alert",
    ]
    title = (
        f"Suggested Matches — {run.deductor_name} | Run #{run.run_number} "
        f"| {run.financial_year}"
    )
    _write_title_header_sep(ws, title, headers, get_column_letter(len(headers)), title_fill=AMBER)

    conf_colors = {"HIGH": CONF_HIGH, "MEDIUM": CONF_MED, "LOW": CONF_LOW}

    for r, s in enumerate(suggestions, 4):
        conf_color = conf_colors.get(s.confidence or "", "FFFFFF")
        var_pct = s.variance_pct or 0
        score = s.composite_score or 0
        var_color = VAR_GREEN if var_pct <= var_green else VAR_YELLOW if var_pct <= var_yellow else VAR_RED
        cat_color = CAT_COLORS.get(s.category, "888888")

        ws.cell(r, 1, r - 3).alignment = _align("center")
        cat_cell = ws.cell(r, 2, (s.category or "—").replace("_", " "))
        cat_cell.fill = _fill(cat_color)
        cat_cell.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        ws.cell(r, 3, s.as26_date or "—")
        ws.cell(r, 4, s.section or "—").alignment = _align("center")
        ws.cell(r, 5, s.as26_amount or 0).number_format = '#,##0.00'
        ws.cell(r, 6, s.books_sum or 0).number_format = '#,##0.00'
        ws.cell(r, 7, s.variance_amt or 0).fill = _fill(var_color); ws.cell(r, 7).number_format = '#,##0.00'
        ws.cell(r, 8, f"{var_pct:.2f}%").fill = _fill(var_color)
        ws.cell(r, 9, s.match_type or "—")
        ws.cell(r, 10, s.confidence or "—").fill = _fill(conf_color)
        ws.cell(r, 11, f"{score:.1f}")
        ws.cell(r, 12, ", ".join(s.invoice_refs or []))
        # Status
        if s.authorized:
            status_text, status_fill = "✓ Authorized", VAR_GREEN
        elif s.rejected:
            status_text, status_fill = "✗ Rejected", VAR_RED
        else:
            status_text, status_fill = "⏳ Pending", VAR_YELLOW
        ws.cell(r, 13, status_text).fill = _fill(status_fill)
        ws.cell(r, 14, s.alert_message or "").alignment = _align("left", wrap=True)
        ws.row_dimensions[r].height = 30

    widths = [5, 22, 14, 10, 16, 16, 14, 12, 18, 11, 14, 35, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_unmatched_26as(ws, entries: List[Unmatched26AS]):
    headers = ["#", "Deductor", "TAN", "Date", "Amount (₹)", "Section",
               "Nearest Invoice", "Nearest Var %", "Reason"]
    _write_title_header_sep(ws, "Unmatched Form 26AS Entries", headers,
                            get_column_letter(len(headers)), title_fill="B71C1C")

    for r, e in enumerate(entries, 4):
        ws.cell(r, 1, r - 3).alignment = _align("center")
        ws.cell(r, 2, e.deductor_name)
        ws.cell(r, 3, e.tan).alignment = _align("center")
        ws.cell(r, 4, e.transaction_date)
        ws.cell(r, 5, e.amount).number_format = '#,##0.00'
        ws.cell(r, 5).fill = _fill(VAR_RED)
        ws.cell(r, 6, e.section).alignment = _align("center")
        ws.cell(r, 7, getattr(e, 'best_candidate_invoice', None) or "—")
        bv = getattr(e, 'best_candidate_variance_pct', None)
        ws.cell(r, 8, f"{bv:.1f}%" if bv is not None else "—").alignment = _align("center")
        ws.cell(r, 9, f"[{e.reason_code}] {e.reason_detail}").alignment = _align("left", wrap=True)
        ws.row_dimensions[r].height = 30

    widths = [5, 35, 14, 14, 16, 10, 20, 12, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_unmatched_books(ws, entries: List[UnmatchedBook]):
    headers = ["#", "Invoice Ref", "Amount (₹)", "Doc Date", "Doc Type", "Clearing Doc", "Flags"]
    _write_title_header_sep(ws, "Unmatched SAP Book Entries", headers,
                            get_column_letter(len(headers)), title_fill="1565C0")

    for r, b in enumerate(entries, 4):
        ws.cell(r, 1, r - 3).alignment = _align("center")
        ws.cell(r, 2, b.invoice_ref)
        ws.cell(r, 3, b.amount).number_format = '#,##0.00'
        ws.cell(r, 4, b.doc_date or "—")
        ws.cell(r, 5, b.doc_type or "—").alignment = _align("center")
        ws.cell(r, 6, b.clearing_doc or "—")
        ws.cell(r, 7, b.flag or "—")

    widths = [5, 25, 16, 14, 12, 16, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Batch Excel ──────────────────────────────────────────────────────────────

def generate_batch_excel(
    runs_data: List[dict],
    variance_thresholds: tuple[float, float] = (1.0, 3.0),
) -> bytes:
    """Generate a combined Excel workbook for a batch of runs.

    Args:
        runs_data: list of dicts, each with keys:
            run, matched_pairs, unmatched_26as, unmatched_books, exceptions,
            suggested_matches (optional)
        variance_thresholds: (green_ceiling, yellow_ceiling) from admin config.
    Returns:
        Excel file bytes.
    """
    wb = Workbook()

    # ── Batch Summary sheet ─────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Batch Summary"
    ws_summary.sheet_view.showGridLines = False
    ws_summary.freeze_panes = "A3"

    first_run = runs_data[0]["run"] if runs_data else None
    fy_label = first_run.financial_year if first_run else "—"

    summary_headers = [
        "#", "Deductor", "TAN", "Status", "Match Rate",
        "Matched", "Unmatched 26AS", "Total 26AS",
        "Violations", "HIGH", "MEDIUM", "LOW",
    ]

    ws_summary.merge_cells("A1:L1")
    ws_summary["A1"] = f"TDS BATCH RECONCILIATION SUMMARY — {fy_label}"
    ws_summary["A1"].font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    ws_summary["A1"].fill = _fill(NAVY)
    ws_summary["A1"].alignment = _align("center")
    ws_summary.row_dimensions[1].height = 30

    for c, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(2, c, h)
        cell.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        cell.fill = _fill(NAVY)
        cell.alignment = _align("center")

    total_matched_all = 0
    total_26as_all = 0

    for idx, rd in enumerate(runs_data, 3):
        run: ReconciliationRun = rd["run"]
        total_matched_all += run.matched_count or 0
        total_26as_all += run.total_26as_entries or 0

        rate = run.match_rate_pct or 0
        rate_color = VAR_GREEN if rate >= 95 else VAR_YELLOW if rate >= 75 else VAR_RED

        ws_summary.cell(idx, 1, idx - 2).alignment = _align("center")
        ws_summary.cell(idx, 2, run.deductor_name or "—")
        ws_summary.cell(idx, 3, run.tan or "—").alignment = _align("center")
        ws_summary.cell(idx, 4, run.status or "—")
        rate_cell = ws_summary.cell(idx, 5, f"{rate:.2f}%")
        rate_cell.fill = _fill(rate_color)
        rate_cell.alignment = _align("center")
        ws_summary.cell(idx, 6, run.matched_count or 0).alignment = _align("center")
        ws_summary.cell(idx, 7, run.unmatched_26as_count or 0).alignment = _align("center")
        ws_summary.cell(idx, 8, run.total_26as_entries or 0).alignment = _align("center")
        viol_cell = ws_summary.cell(idx, 9, run.constraint_violations or 0)
        viol_cell.alignment = _align("center")
        if (run.constraint_violations or 0) > 0:
            viol_cell.fill = _fill(VAR_RED)
        ws_summary.cell(idx, 10, run.high_confidence_count or 0).alignment = _align("center")
        ws_summary.cell(idx, 10).fill = _fill(CONF_HIGH)
        ws_summary.cell(idx, 11, run.medium_confidence_count or 0).alignment = _align("center")
        ws_summary.cell(idx, 11).fill = _fill(CONF_MED)
        ws_summary.cell(idx, 12, run.low_confidence_count or 0).alignment = _align("center")
        ws_summary.cell(idx, 12).fill = _fill(CONF_LOW)

    # Totals row
    totals_row = len(runs_data) + 3
    ws_summary.cell(totals_row, 1, "").font = Font(bold=True, size=9, name="Calibri")
    ws_summary.cell(totals_row, 2, "TOTAL").font = Font(bold=True, size=10, name="Calibri")
    overall_rate = (total_matched_all / total_26as_all * 100) if total_26as_all > 0 else 0
    rate_color = VAR_GREEN if overall_rate >= 95 else VAR_YELLOW if overall_rate >= 75 else VAR_RED
    rate_cell = ws_summary.cell(totals_row, 5, f"{overall_rate:.2f}%")
    rate_cell.font = Font(bold=True, size=10, name="Calibri")
    rate_cell.fill = _fill(rate_color)
    rate_cell.alignment = _align("center")
    ws_summary.cell(totals_row, 6, total_matched_all).font = Font(bold=True, size=10, name="Calibri")
    ws_summary.cell(totals_row, 6).alignment = _align("center")
    total_unmatched = sum((rd["run"].unmatched_26as_count or 0) for rd in runs_data)
    ws_summary.cell(totals_row, 7, total_unmatched).font = Font(bold=True, size=10, name="Calibri")
    ws_summary.cell(totals_row, 7).alignment = _align("center")
    ws_summary.cell(totals_row, 8, total_26as_all).font = Font(bold=True, size=10, name="Calibri")
    ws_summary.cell(totals_row, 8).alignment = _align("center")
    total_violations = sum((rd["run"].constraint_violations or 0) for rd in runs_data)
    ws_summary.cell(totals_row, 9, total_violations).font = Font(bold=True, size=10, name="Calibri")
    ws_summary.cell(totals_row, 9).alignment = _align("center")

    # Timestamp
    ts_row = totals_row + 2
    ws_summary.merge_cells(f"A{ts_row}:L{ts_row}")
    ws_summary.cell(ts_row, 1,
                    f"Generated: {datetime.now(timezone.utc).strftime('%d-%b-%Y %H:%M UTC')} | "
                    f"Parties: {len(runs_data)} | Algorithm: {first_run.algorithm_version if first_run else '—'}"
                    ).font = Font(italic=True, size=8, color="888888", name="Calibri")

    summary_widths = [5, 35, 14, 16, 12, 10, 14, 10, 10, 8, 10, 8]
    for i, w in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # ── Per-party sheets ─────────────────────────────────────────────────────
    var_green = variance_thresholds[0]
    var_yellow = variance_thresholds[1]

    for rd in runs_data:
        run: ReconciliationRun = rd["run"]
        # Deduplicate matched pairs by as26_row_hash (safety net against duplicate promotions)
        raw_pairs: List[MatchedPair] = rd["matched_pairs"]
        seen_hashes: set = set()
        matched_pairs: List[MatchedPair] = []
        for mp in raw_pairs:
            h = mp.as26_row_hash
            if h and h in seen_hashes:
                continue
            if h:
                seen_hashes.add(h)
            matched_pairs.append(mp)

        # Excel sheet names max 31 chars, must be unique, no invalid chars
        raw_name = run.deductor_name or run.tan or f"Run-{run.run_number}"
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', raw_name)
        sheet_name = safe_name[:31]
        # Ensure uniqueness
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            suffix = f" ({run.run_number})"
            sheet_name = safe_name[:31 - len(suffix)] + suffix

        ws = wb.create_sheet(sheet_name)
        _build_matched(ws, matched_pairs, run, var_green=var_green, var_yellow=var_yellow)

        # Suggested matches sheet (only if any exist)
        suggested: List[SuggestedMatch] = rd.get("suggested_matches", [])
        if suggested:
            sug_sheet_name = f"Review-{safe_name[:20]}"
            existing = [ws.title for ws in wb.worksheets]
            if sug_sheet_name in existing:
                sug_sheet_name = f"Review-{safe_name[:14]} ({run.run_number})"
            ws_sug = wb.create_sheet(sug_sheet_name)
            _build_suggested(ws_sug, suggested, run, var_green=var_green, var_yellow=var_yellow)

        # Unmatched 26AS sheet (only if any exist)
        unmatched_26as: List[Unmatched26AS] = rd.get("unmatched_26as", [])
        if unmatched_26as:
            un_sheet_name = f"Unmatched-{safe_name[:18]}"
            existing = [ws.title for ws in wb.worksheets]
            if un_sheet_name in existing:
                un_sheet_name = f"Unmatched-{safe_name[:12]} ({run.run_number})"
            ws_un = wb.create_sheet(un_sheet_name)
            _build_unmatched_26as(ws_un, unmatched_26as)

        # Exceptions sheet (only if any exist)
        exceptions: List[ExceptionRecord] = rd.get("exceptions", [])
        if exceptions:
            exc_sheet_name = f"Exceptions-{safe_name[:16]}"
            existing = [ws.title for ws in wb.worksheets]
            if exc_sheet_name in existing:
                exc_sheet_name = f"Exceptions-{safe_name[:10]} ({run.run_number})"
            ws_exc = wb.create_sheet(exc_sheet_name)
            _build_exceptions(ws_exc, exceptions)

        # Unmatched SAP Books sheet (only if any exist)
        unmatched_books: List[UnmatchedBook] = rd.get("unmatched_books", [])
        if unmatched_books:
            bks_sheet_name = f"SAP-Unmtchd-{safe_name[:16]}"
            existing = [ws.title for ws in wb.worksheets]
            if bks_sheet_name in existing:
                bks_sheet_name = f"SAP-Unmtchd-{safe_name[:10]} ({run.run_number})"
            ws_bks = wb.create_sheet(bks_sheet_name)
            _build_unmatched_books(ws_bks, unmatched_books)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_variance(ws, pairs: List[MatchedPair]):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Variance Analysis"
    ws["A1"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    ws["A1"].fill = _fill(PURPLE)
    ws["A1"].alignment = _align("center")

    # Distribution buckets
    buckets = {"0-1%": 0, "1-2%": 0, "2-3%": 0, "3-5%": 0, ">5%": 0, "Exact (₹0.01)": 0}
    for p in pairs:
        v = p.variance_pct
        if v <= 0.01:
            buckets["Exact (₹0.01)"] += 1
        elif v <= 1.0:
            buckets["0-1%"] += 1
        elif v <= 2.0:
            buckets["1-2%"] += 1
        elif v <= 3.0:
            buckets["2-3%"] += 1
        elif v <= 5.0:
            buckets["3-5%"] += 1
        else:
            buckets[">5%"] += 1

    ws.cell(2, 1, "Variance Band").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(2, 2, "Count").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(2, 3, "% of Matched").font = Font(bold=True, size=10, name="Calibri")

    total = len(pairs)
    for i, (band, count) in enumerate(buckets.items(), 3):
        ws.cell(i, 1, band)
        ws.cell(i, 2, count)
        ws.cell(i, 3, f"{count/total*100:.1f}%" if total > 0 else "—")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
