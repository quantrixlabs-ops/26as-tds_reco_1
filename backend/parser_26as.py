"""
26AS Master File Parser — Phase 1
Pure function: parse_26as(file_bytes) → as26_df

Handles two layouts automatically:
  ① Full master 26AS  — Row 1 blank, Row 2 grand totals, Row 3 headers, Row 4+ data
  ② Single-deductor slice — Row 1 headers, Row 2+ data  (as in provided sample)

Only rows with Status of Booking = 'F' (Final) are returned.
Match column: 'Amount Paid / Credited(Rs.)' — NOT Tax Deducted or TDS Deposited.

Column names (header-detected, not positional — 26AS layout varies by portal version):
  Name of Deductor        → deductor_name
  TAN of Deductor         → tan
  Section                 → section
  Transaction Date        → transaction_date
  Status of Booking       → status
  Amount Paid/Credited    → amount
  Invoice Number          → invoice_number
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Any, Optional

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# Canonical column name normaliser
# Expanded aliases to handle various 26AS portal formats (FY20-21 through FY25-26)
_AMOUNT_ALIASES = re.compile(r"amount\s*paid|amount\s*credited|^amount$", re.IGNORECASE)
_NAME_ALIASES   = re.compile(r"name\s+of\s+deductor|^particulars$|deductor\s*name|^name$", re.IGNORECASE)
_TAN_ALIASES    = re.compile(r"tan\s+of\s+deductor|^tan$", re.IGNORECASE)
_STATUS_ALIASES = re.compile(r"status\s+of\s+booking|^status\s+of\b", re.IGNORECASE)
_DATE_ALIASES   = re.compile(r"transaction\s+date|^date$|date\s+of\s+(payment|credit)", re.IGNORECASE)
_SECTION_ALIAS  = re.compile(r"^section$", re.IGNORECASE)
_INVOICE_ALIAS  = re.compile(r"invoice\s+number|^invoice\s*no", re.IGNORECASE)
_TDS_AMT_ALIAS  = re.compile(r"tax\s+deducted|tds\s+deducted|^tds$|tax\s+amount", re.IGNORECASE)


def _normalise_col(name: str) -> Optional[str]:
    s = str(name).strip()
    if _AMOUNT_ALIASES.search(s):   return "amount"
    if _NAME_ALIASES.search(s):     return "deductor_name"
    if _TAN_ALIASES.search(s):      return "tan"
    if _STATUS_ALIASES.search(s):   return "status"
    if _DATE_ALIASES.search(s):     return "transaction_date"
    if _SECTION_ALIAS.match(s):     return "section"
    if _INVOICE_ALIAS.search(s):    return "invoice_number"
    if _TDS_AMT_ALIAS.search(s):    return "tds_amount"
    return None


def _parse_date(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d-%b-%Y")
    if isinstance(val, date):
        return val.strftime("%d-%b-%Y")
    s = str(val).strip()
    if not s:
        return None
    # Try multiple formats before giving up
    for fmt in [None, "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"]:
        try:
            if fmt:
                return datetime.strptime(s, fmt).strftime("%d-%b-%Y")
            else:
                return pd.to_datetime(s, dayfirst=True).strftime("%d-%b-%Y")
        except Exception:
            continue
    return None  # Reject unparseable dates instead of returning garbage strings


def _detect_header_row(ws, lenient: bool = False) -> int:
    """
    Return the 1-based row number that contains the 26AS column headers.
    Looks for 'Name of Deductor' (or 'Amount Paid' in lenient mode) within first 5 rows.
    Falls back to row 3 (spec default) then row 1.
    """
    def _row_vals(row_num: int):
        try:
            return [c.value for c in next(ws.iter_rows(min_row=row_num, max_row=row_num))]
        except StopIteration:
            return []

    # Primary: look for deductor name header
    for row_num in range(1, 6):
        for val in _row_vals(row_num):
            if val and _NAME_ALIASES.search(str(val)):
                return row_num
    # Lenient fallback: look for amount or status header (files without deductor info)
    if lenient:
        for row_num in range(1, 6):
            for val in _row_vals(row_num):
                if val and (_AMOUNT_ALIASES.search(str(val)) or _STATUS_ALIASES.search(str(val))):
                    return row_num
    # Spec default
    return 3


def _parse_date_to_date(val: Any) -> Optional[date]:
    """Return a date object from a cell value, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return pd.to_datetime(str(val), dayfirst=True).date()
    except Exception:
        return None


def parse_26as(
    file_bytes: bytes,
    fy_start: Optional[date] = None,
    fy_end: Optional[date] = None,
    lenient: bool = False,
) -> pd.DataFrame:
    """
    Parse a 26AS Excel file and return a DataFrame with Status=F rows only.

    Parameters
    ----------
    file_bytes : Raw .xlsx bytes
    fy_start   : If provided, exclude rows where transaction_date < fy_start
    fy_end     : If provided, exclude rows where transaction_date > fy_end
    lenient    : If True, deductor_name and tan are optional (filled with
                 "UNKNOWN" / "UNKNOWN_TAN"). Useful for single-party 26AS
                 files that don't have deductor info columns.

    Columns returned:
        deductor_name, tan, section, transaction_date, amount,
        status, invoice_number
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # ── Detect which sheet holds transaction data ──────────────────────────
    data_sheet = None
    for shname in wb.sheetnames:
        if "tanwise" not in shname.lower() and "summary" not in shname.lower():
            data_sheet = wb[shname]
            break
    if data_sheet is None:
        data_sheet = wb.active

    # ── Detect header row ──────────────────────────────────────────────────
    header_row_num = _detect_header_row(data_sheet, lenient=lenient)
    data_start_row = header_row_num + 1

    # ── Read header row and build column mapping ───────────────────────────
    header_cells = list(data_sheet.iter_rows(
        min_row=header_row_num, max_row=header_row_num, values_only=True
    ))[0]

    col_map: dict[int, str] = {}  # 0-based index → canonical name
    used_canonicals: set[str] = set()  # Prevent duplicate canonical names
    for idx, cell_val in enumerate(header_cells):
        if cell_val is None:
            continue
        canonical = _normalise_col(str(cell_val))
        if canonical and canonical not in used_canonicals:
            col_map[idx] = canonical
            used_canonicals.add(canonical)

    required = {"amount", "status"} if lenient else {"deductor_name", "tan", "amount", "status"}
    found = set(col_map.values())
    if not required.issubset(found):
        missing = required - found
        raise ValueError(
            f"26AS file is missing required columns: {missing}. "
            f"Found headers: {[h for h in header_cells if h]}"
        )

    # ── Read data rows ─────────────────────────────────────────────────────
    rows = []
    non_f_count = 0
    date_excluded_count = 0
    for raw_row in data_sheet.iter_rows(min_row=data_start_row, values_only=True):
        # Skip entirely blank rows
        if all(v is None for v in raw_row):
            continue

        mapped: dict[str, Any] = {v: None for v in col_map.values()}
        for idx, canonical in col_map.items():
            if idx < len(raw_row):
                mapped[canonical] = raw_row[idx]

        # Filter: Status = 'F' only
        status_val = str(mapped.get("status") or "").strip().upper()
        if status_val != "F":
            non_f_count += 1
            continue

        # Parse amount
        raw_amt = mapped.get("amount")
        if raw_amt is None:
            continue
        try:
            amount = float(raw_amt)
        except (ValueError, TypeError):
            continue
        if amount == 0:
            continue

        # ── FY date-range filter ──────────────────────────────────────────
        if fy_start and fy_end:
            txn_date = _parse_date_to_date(mapped.get("transaction_date"))
            if txn_date is None:
                # Exclude rows with unparseable dates when FY filter is active
                date_excluded_count += 1
                continue
            if not (fy_start <= txn_date <= fy_end):
                date_excluded_count += 1
                continue

        # Parse TDS amount (optional — activates rate validator when present)
        raw_tds = mapped.get("tds_amount")
        tds_amount = None
        if raw_tds is not None:
            try:
                tds_amount = float(raw_tds)
            except (ValueError, TypeError):
                pass

        rows.append({
            "deductor_name":    str(mapped.get("deductor_name") or "").strip(),
            "tan":              str(mapped.get("tan") or "").strip().upper(),
            "section":          str(mapped.get("section") or "").strip(),
            "transaction_date": _parse_date(mapped.get("transaction_date")),
            "amount":           amount,
            "tds_amount":       tds_amount,
            "status":           "F",
            "invoice_number":   str(mapped.get("invoice_number") or "").strip()
                                if mapped.get("invoice_number") else "",
        })

    wb.close()

    df = pd.DataFrame(rows)
    if df.empty:
        logger.warning("26AS parse returned 0 rows with Status=F")
        return pd.DataFrame(columns=[
            "deductor_name", "tan", "section", "transaction_date",
            "amount", "tds_amount", "status", "invoice_number"
        ])

    if non_f_count > 0:
        logger.info("26AS: excluded %d rows with status != F", non_f_count)
    if date_excluded_count > 0:
        logger.info("26AS: excluded %d rows by FY date filter (incl. unparseable dates)", date_excluded_count)
    logger.info(
        "26AS parse: %d rows (Status=F) | deductors: %d | header at row %d",
        len(df), df["deductor_name"].nunique(), header_row_num,
    )
    return df


def get_tanwise_candidates(file_bytes: bytes) -> list[dict]:
    """
    Extract name+TAN pairs from TANWISE SUMMARY sheet (if present).
    Returns list of {deductor_name, tan}.
    Used as additional candidate pool for fuzzy matching.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    candidates = []

    tanwise_sheet = None
    for shname in wb.sheetnames:
        if "tanwise" in shname.lower() or ("tan" in shname.lower() and "summary" in shname.lower()):
            tanwise_sheet = wb[shname]
            break

    if tanwise_sheet is None:
        wb.close()
        return candidates

    # Find header row — look for TAN column
    header_row = None
    name_col_idx = None
    tan_col_idx = None

    for row_num in range(1, 6):
        row_vals = [c.value for c in next(tanwise_sheet.iter_rows(min_row=row_num, max_row=row_num))]
        for idx, v in enumerate(row_vals):
            if v and re.search(r"customer\s+name|deductor", str(v), re.IGNORECASE):
                name_col_idx = idx
            if v and re.search(r"^tan$", str(v).strip(), re.IGNORECASE):
                tan_col_idx = idx
        if name_col_idx is not None and tan_col_idx is not None:
            header_row = row_num
            break

    if header_row is None:
        wb.close()
        return candidates

    for raw_row in tanwise_sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in raw_row):
            continue
        name_val = raw_row[name_col_idx] if name_col_idx < len(raw_row) else None
        tan_val  = raw_row[tan_col_idx]  if tan_col_idx  < len(raw_row) else None
        if name_val and tan_val:
            candidates.append({
                "deductor_name": str(name_val).strip(),
                "tan": str(tan_val).strip().upper(),
            })

    wb.close()
    return candidates
